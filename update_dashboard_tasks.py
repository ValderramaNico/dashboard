import argparse
import datetime
import json
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_EXCEL_PATH = PROJECT_ROOT / "documentos" / "avangrid" / "Native Mobile - CMP Regression Automation.xlsx"
DEFAULT_JSON_PATH = PROJECT_ROOT / "public" / "daily_tasks.json"
DEFAULT_HISTORY_PATH = PROJECT_ROOT / "public" / "task_history.json"


def parse_args():
    parser = argparse.ArgumentParser(description="Exporta las tareas QA marcadas para una fecha a JSON.")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH, help="Ruta al libro de Excel.")
    parser.add_argument("--output", type=Path, default=DEFAULT_JSON_PATH, help="Archivo JSON de salida.")
    parser.add_argument("--history", type=Path, default=DEFAULT_HISTORY_PATH, help="Archivo JSON con el historial por día.")
    parser.add_argument("--date", type=datetime.date.fromisoformat, default=datetime.date.today(), help="Fecha a procesar (AAAA-MM-DD).")
    return parser.parse_args()

def main():
    args = parse_args()
    excel_path = args.excel
    json_output_path = args.output
    history_output_path = args.history

    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    # Get current date
    today = args.date
    day_num_str = str(today.day)
    month_name = today.strftime("%B")  # e.g. "July"
    
    print(f"Scanning Excel for date: {today} ({month_name} {day_num_str})")
    
    sheets_to_scan = ['ManagePayments', 'MakePayment', 'Login', 'Autopay', 'Outages']
    today_tasks = []

    for name in sheets_to_scan:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        
        # 1. Find the column for today
        today_col = None
        max_c = ws.max_column
        for c in range(1, max_c + 1):
            # Resolve month name for column c
            col_month = None
            for temp_c in range(c, 0, -1):
                m_val = ws.cell(row=1, column=temp_c).value
                if m_val is not None:
                    col_month = str(m_val).strip()
                    break
            
            col_day_num = ws.cell(row=3, column=c).value
            
            # Match current month and day number
            if col_month == month_name and (str(col_day_num) == day_num_str or col_day_num == today.day):
                today_col = c
                break
                
        if today_col is None:
            print(f"  Info: No column found for {month_name} {day_num_str} in sheet '{name}'")
            continue
            
        print(f"  Scanning sheet '{name}', found today's column at index {today_col}")
        
        # 2. Iterate rows starting from row 5
        for r in range(5, ws.max_row + 1):
            status_val = ws.cell(row=r, column=4).value
            if status_val == "QA Validation":
                cell = ws.cell(row=r, column=today_col)
                fill = cell.fill
                
                # Check for background fill (excluding default weekend pattern)
                has_fill = False
                color_val = None
                if fill and fill.fill_type and fill.fill_type != 'none':
                    # We check if fgColor is valid
                    if fill.fgColor:
                        color_val = fill.fgColor.value
                        # Color '00000000' or 0 is default / no fill in some styles, 
                        # but custom colors will have hex values like 'FFFFC000' (orange)
                        if color_val and str(color_val) not in ['00000000', '0', '000000']:
                            has_fill = True
                
                if has_fill:
                    tc_name = ws.cell(row=r, column=1).value
                    tc_key = ws.cell(row=r, column=2).value
                    assigned = ws.cell(row=r, column=3).value
                    
                    today_tasks.append({
                        "date": today.strftime("%Y-%m-%d"),
                        "sheet": name,
                        "key": tc_key,
                        "name": tc_name,
                        "assigned": assigned
                    })

    # Save the latest snapshot for backwards compatibility.
    json_output_path.parent.mkdir(parents=True, exist_ok=True)
    with json_output_path.open("w", encoding="utf-8") as f:
        json.dump(today_tasks, f, indent=2, ensure_ascii=False)

    # Keep all previous dates. Re-running a date replaces only that day's snapshot.
    history = {"version": 1, "days": {}}
    if history_output_path.is_file():
        with history_output_path.open("r", encoding="utf-8") as f:
            existing_history = json.load(f)
        if isinstance(existing_history, dict) and isinstance(existing_history.get("days"), dict):
            history = existing_history

    history["version"] = 1
    history["updatedAt"] = datetime.datetime.now(datetime.timezone.utc).isoformat()
    history["days"][today.isoformat()] = today_tasks

    history_output_path.parent.mkdir(parents=True, exist_ok=True)
    with history_output_path.open("w", encoding="utf-8") as f:
        json.dump(history, f, indent=2, ensure_ascii=False)

    print(f"\nSuccessfully found {len(today_tasks)} tasks. Saved latest snapshot to {json_output_path} and history to {history_output_path}")

if __name__ == "__main__":
    main()
